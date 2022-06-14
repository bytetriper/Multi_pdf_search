from ast import keyword
from cgi import print_directory
from enum import Flag
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import filesearch
import os
files=filesearch.filesearch(r'python\openpyxl',typename='xlsx',output=True)
def searchexcel(key,filename):
    exc=load_workbook(filename=filename)
    sheets=exc.sheetnames
    fg=False
    for sheet in sheets:
        worksheet=exc[sheet]
        i=0
        j=0
        #print(sheet)
        for row in worksheet.iter_rows(values_only=True):
            i+=1
            for value in row:
                j+=1
                if str(key)==str(value):
                    fg=True
                    print("find in file "+filename+"-sheet "+sheet+"-row "+str(i)+":")
                    print(row)
                    break
    if fg==False:
        print("No match in file "+filename)
    return
keyword=''
while 1:
    keyword=input("Search Keywords:")
    for i in files:
        searchexcel(keyword,'python\\openpyxl\\'+i)